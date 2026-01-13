#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
测试版本一致性检查功能
演示如何检查本地Excel文件是否基于远程最新版本
"""

import os
import shutil
from datetime import datetime
from openpyxl import load_workbook

EXCEL_FILE = "excels/数据文件_001.xlsx"
BACKUP_FILE = EXCEL_FILE + ".backup"

print("=" * 60)
print("测试版本一致性检查功能")
print("=" * 60)

# 备份文件
if os.path.exists(EXCEL_FILE):
    shutil.copy(EXCEL_FILE, BACKUP_FILE)
    print(f"已备份文件: {BACKUP_FILE}")

# 场景1: 模拟远程有新修订，本地未更新
print("\n场景1: 模拟远程有新修订，本地未更新")
print("-" * 60)

# 先提交当前版本到远程（模拟远程版本）
print("步骤1: 提交当前版本到远程")
print("执行: git add excels/数据文件_001.xlsx")
print("      git commit -m '提交当前版本'")
print("      git push")
print("\n按回车继续...")
input()

# 模拟远程添加新修订记录
print("\n步骤2: 模拟远程添加新修订记录")
wb = load_workbook(EXCEL_FILE)
ws_revision = wb["修改记录"]
last_row = ws_revision.max_row
revision_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
ws_revision.cell(row=last_row + 1, column=1, value="远程用户")
ws_revision.cell(row=last_row + 1, column=2, value=revision_time)
ws_revision.cell(row=last_row + 1, column=3, value="远程修改：添加了新数据")
ws_revision.cell(row=last_row + 1, column=4, value=f"v{last_row}.0")
wb.save(EXCEL_FILE)
wb.close()

print(f"已添加远程修订记录: 远程用户 - {revision_time}")
print("执行: git add excels/数据文件_001.xlsx")
print("      git commit -m '远程添加新修订'")
print("      git push")
print("\n按回车继续...")
input()

# 恢复到旧版本（模拟本地未更新）
print("\n步骤3: 恢复到旧版本（模拟本地未更新）")
if os.path.exists(BACKUP_FILE):
    shutil.copy(BACKUP_FILE, EXCEL_FILE)
    print(f"已恢复到旧版本: {EXCEL_FILE}")

# 本地修改但不包含远程的新修订记录
print("\n步骤4: 本地修改但不包含远程的新修订记录")
wb = load_workbook(EXCEL_FILE)
ws = wb["数据表1"]
ws.cell(row=2, column=3, value=99999)

# 添加本地修订记录（但不包含远程的新修订）
ws_revision = wb["修改记录"]
last_row = ws_revision.max_row
local_revision_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
ws_revision.cell(row=last_row + 1, column=1, value="本地用户")
ws_revision.cell(row=last_row + 1, column=2, value=local_revision_time)
ws_revision.cell(row=last_row + 1, column=3, value="本地修改：更新了数据")
ws_revision.cell(row=last_row + 1, column=4, value=f"v{last_row}.0")

wb.save(EXCEL_FILE)
wb.close()

print(f"已添加本地修订记录: 本地用户 - {local_revision_time}")
print("\n现在执行以下命令测试:")
print("  git add excels/数据文件_001.xlsx")
print("  git commit -m '本地修改'")
print("\n预期结果: 提交被拦截，提示'本地文件未基于远程最新版本'")
print("\n按回车继续...")
input()

# 场景2: 先pull再修改
print("\n场景2: 先pull再修改")
print("-" * 60)

print("步骤1: 先执行 git pull 获取最新版本")
print("执行: git pull")
print("\n按回车继续...")
input()

print("\n步骤2: 在最新版本基础上修改")
wb = load_workbook(EXCEL_FILE)
ws = wb["数据表1"]
ws.cell(row=2, column=3, value=88888)

# 添加修订记录（包含远程的新修订）
ws_revision = wb["修改记录"]
last_row = ws_revision.max_row
new_revision_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
ws_revision.cell(row=last_row + 1, column=1, value="本地用户")
ws_revision.cell(row=last_row + 1, column=2, value=new_revision_time)
ws_revision.cell(row=last_row + 1, column=3, value="本地修改：在最新版本基础上更新")
ws_revision.cell(row=last_row + 1, column=4, value=f"v{last_row}.0")

wb.save(EXCEL_FILE)
wb.close()

print(f"已添加修订记录: 本地用户 - {new_revision_time}")
print("\n现在执行以下命令测试:")
print("  git add excels/数据文件_001.xlsx")
print("  git commit -m '在最新版本基础上修改'")
print("\n预期结果: 检查通过，提交成功")
print("\n按回车继续...")
input()

# 恢复文件
print("\n恢复原始文件...")
if os.path.exists(BACKUP_FILE):
    shutil.copy(BACKUP_FILE, EXCEL_FILE)
    print(f"已恢复文件: {EXCEL_FILE}")

print("\n测试完成！")
print("\n总结:")
print("1. 如果本地文件未包含远程最新的修订记录，提交会被拦截")
print("2. 需要先执行 git pull 获取最新版本")
print("3. 在最新版本基础上进行修改并添加修订记录")
print("4. 这样可以避免覆盖他人的修改")
