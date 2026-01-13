#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel检查功能测试脚本
演示如何测试Excel文件的检查功能
"""

import os
import shutil
from datetime import datetime
from openpyxl import load_workbook

# 常量定义
EXCEL_DIR = "excels"
TEST_FILE = "数据文件_001.xlsx"
TEST_FILE_PATH = os.path.join(EXCEL_DIR, TEST_FILE)
BACKUP_FILE = TEST_FILE_PATH + ".backup"


def backup_file():
    """备份测试文件"""
    if os.path.exists(TEST_FILE_PATH):
        shutil.copy(TEST_FILE_PATH, BACKUP_FILE)
        print(f"已备份文件: {BACKUP_FILE}")
        return True
    return False


def restore_file():
    """恢复测试文件"""
    if os.path.exists(BACKUP_FILE):
        shutil.copy(BACKUP_FILE, TEST_FILE_PATH)
        print(f"已恢复文件: {TEST_FILE_PATH}")
        return True
    return False


def modify_excel_without_revision():
    """修改Excel文件但不添加修订记录"""
    print("\n" + "=" * 60)
    print("测试场景1: 修改Excel但不添加修订记录")
    print("=" * 60)
    
    wb = load_workbook(TEST_FILE_PATH)
    
    # 修改数据表1的第一个单元格
    ws = wb["数据表1"]
    ws.cell(row=2, column=3, value=99999)
    
    wb.save(TEST_FILE_PATH)
    wb.close()
    
    print(f"已修改 {TEST_FILE} 的数据表1")
    print("未添加修订记录")
    print("\n现在执行以下命令测试:")
    print(f"  git add {TEST_FILE_PATH}")
    print(f"  git commit -m '修改数据文件'")
    print("\n预期结果: 提交被拦截，提示'修改记录sheet页为空'")


def modify_excel_with_revision():
    """修改Excel文件并添加修订记录"""
    print("\n" + "=" * 60)
    print("测试场景2: 修改Excel并添加修订记录")
    print("=" * 60)
    
    wb = load_workbook(TEST_FILE_PATH)
    
    # 修改数据表1的第一个单元格
    ws = wb["数据表1"]
    ws.cell(row=2, column=3, value=88888)
    
    # 添加修订记录
    ws_revision = wb["修改记录"]
    
    # 找到最后一行
    last_row = ws_revision.max_row
    
    # 添加新修订记录
    revision_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws_revision.cell(row=last_row + 1, column=1, value="测试用户")
    ws_revision.cell(row=last_row + 1, column=2, value=revision_time)
    ws_revision.cell(row=last_row + 1, column=3, value="测试修改：更新了数据")
    ws_revision.cell(row=last_row + 1, column=4, value=f"v{last_row}.0")
    
    wb.save(TEST_FILE_PATH)
    wb.close()
    
    print(f"已修改 {TEST_FILE} 的数据表1")
    print(f"已添加修订记录: 测试用户 - {revision_time}")
    print("\n现在执行以下命令测试:")
    print(f"  git add {TEST_FILE_PATH}")
    print(f"  git commit -m '修改数据文件并添加修订记录'")
    print("\n预期结果: 检查通过，提交成功")


def clear_revision_records():
    """清空修订记录"""
    print("\n" + "=" * 60)
    print("测试场景3: 清空修订记录")
    print("=" * 60)
    
    wb = load_workbook(TEST_FILE_PATH)
    ws_revision = wb["修改记录"]
    
    # 删除所有数据行（保留表头）
    ws_revision.delete_rows(2, ws_revision.max_row - 1)
    
    wb.save(TEST_FILE_PATH)
    wb.close()
    
    print(f"已清空 {TEST_FILE} 的修订记录")
    print("\n现在执行以下命令测试:")
    print(f"  git add {TEST_FILE_PATH}")
    print(f"  git commit -m '清空修订记录'")
    print("\n预期结果: 提交被拦截，提示'修改记录sheet页为空'")


def main():
    """主函数"""
    print("=" * 60)
    print("Excel检查功能测试")
    print("=" * 60)
    
    # 备份文件
    if not backup_file():
        print(f"错误: 找不到测试文件 {TEST_FILE_PATH}")
        return
    
    print("\n请选择测试场景:")
    print("1. 修改Excel但不添加修订记录（应该被拦截）")
    print("2. 修改Excel并添加修订记录（应该通过）")
    print("3. 清空修订记录（应该被拦截）")
    print("4. 恢复原始文件")
    print("0. 退出")
    
    while True:
        choice = input("\n请输入选项 (0-4): ").strip()
        
        if choice == "1":
            modify_excel_without_revision()
        elif choice == "2":
            modify_excel_with_revision()
        elif choice == "3":
            clear_revision_records()
        elif choice == "4":
            restore_file()
        elif choice == "0":
            print("\n退出测试")
            break
        else:
            print("无效选项，请重新输入")


if __name__ == "__main__":
    main()
