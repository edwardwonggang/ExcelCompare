#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel文件生成器
生成100个Excel文件，每个文件包含多个sheet和修改记录
"""

import os
import random
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# 常量定义
EXCEL_COUNT = 100
SHEET_COUNT = 3
RECORD_SHEET_NAME = "修改记录"
OUTPUT_DIR = "excels"

# 修订人列表
REVISERS = ["张三", "李四", "王五", "赵六", "钱七"]

# 修订内容模板
REVISION_CONTENTS = [
    "更新了数据统计",
    "修正了公式错误",
    "添加了新的分析维度",
    "优化了表格格式",
    "补充了缺失数据",
    "调整了计算逻辑",
    "更新了图表数据",
    "修正了拼写错误"
]


def create_header_style():
    """创建表头样式"""
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    return {
        'font': header_font,
        'fill': header_fill,
        'alignment': header_alignment
    }


def create_data_style():
    """创建数据样式"""
    data_font = Font(name='微软雅黑', size=10)
    data_alignment = Alignment(horizontal='left', vertical='center')
    return {
        'font': data_font,
        'alignment': data_alignment
    }


def create_revision_sheet(ws):
    """创建修改记录sheet"""
    headers = ["修订人", "修订时间", "修订内容", "修订版本"]
    
    # 设置表头
    header_style = create_header_style()
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_style['font']
        cell.fill = header_style['fill']
        cell.alignment = header_style['alignment']
    
    # 设置列宽
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 12
    
    # 生成随机修订记录（1-5条）
    record_count = random.randint(1, 5)
    data_style = create_data_style()
    
    for row_idx in range(2, record_count + 2):
        reviser = random.choice(REVISERS)
        # 生成过去30天内的随机时间
        days_ago = random.randint(1, 30)
        revision_time = (datetime.now() - timedelta(days=days_ago)).strftime("%Y-%m-%d %H:%M:%S")
        content = random.choice(REVISION_CONTENTS)
        version = f"v{row_idx - 1}.0"
        
        ws.cell(row=row_idx, column=1, value=reviser)
        ws.cell(row=row_idx, column=2, value=revision_time)
        ws.cell(row=row_idx, column=3, value=content)
        ws.cell(row=row_idx, column=4, value=version)
        
        # 应用数据样式
        for col_idx in range(1, 5):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_style['font']
            cell.alignment = data_style['alignment']
    
    # 冻结首行
    ws.freeze_panes = 'A2'


def create_data_sheet(ws, sheet_name):
    """创建数据sheet"""
    # 生成示例数据
    headers = ["序号", "项目名称", "数值", "状态", "备注"]
    
    # 设置表头
    header_style = create_header_style()
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_style['font']
        cell.fill = header_style['fill']
        cell.alignment = header_style['alignment']
    
    # 设置列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 30
    
    # 生成随机数据（10-20行）
    data_style = create_data_style()
    row_count = random.randint(10, 20)
    
    for row_idx in range(2, row_count + 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=f"{sheet_name}-项目{row_idx - 1}")
        ws.cell(row=row_idx, column=3, value=random.randint(100, 10000))
        ws.cell(row=row_idx, column=4, value=random.choice(["进行中", "已完成", "待处理"]))
        ws.cell(row=row_idx, column=5, value=f"这是{sheet_name}的备注信息{row_idx - 1}")
        
        # 应用数据样式
        for col_idx in range(1, 6):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_style['font']
            cell.alignment = data_style['alignment']
    
    # 冻结首行
    ws.freeze_panes = 'A2'


def generate_excel_files():
    """生成Excel文件"""
    # 创建输出目录
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
    
    print(f"开始生成 {EXCEL_COUNT} 个Excel文件...")
    
    for i in range(1, EXCEL_COUNT + 1):
        wb = Workbook()
        
        # 删除默认sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        
        # 创建修改记录sheet
        ws_revision = wb.create_sheet(RECORD_SHEET_NAME, 0)
        create_revision_sheet(ws_revision)
        
        # 创建数据sheet
        for j in range(1, SHEET_COUNT + 1):
            sheet_name = f"数据表{j}"
            ws_data = wb.create_sheet(sheet_name, j)
            create_data_sheet(ws_data, sheet_name)
        
        # 保存文件
        filename = f"数据文件_{i:03d}.xlsx"
        filepath = os.path.join(OUTPUT_DIR, filename)
        wb.save(filepath)
        
        if i % 10 == 0:
            print(f"已生成 {i} 个文件...")
    
    print(f"✓ 成功生成 {EXCEL_COUNT} 个Excel文件到 {OUTPUT_DIR} 目录")


if __name__ == "__main__":
    generate_excel_files()
