#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel文件检查器
检查Excel文件的版本一致性和修订记录
"""

import os
import sys
import hashlib
import json
import subprocess
import tempfile
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook

# 常量定义
EXCEL_DIR = "excels"
RECORD_SHEET_NAME = "修改记录"
CACHE_FILE = ".excel_cache.json"
CONFIG_FILE = "config.json"

# 默认配置
DEFAULT_CONFIG = {
    "sheet_name": "修改记录",
    "check_columns": ["修订人", "修订时间", "修订内容"],
    "max_threads": 10,
    "timeout": 30
}


class ExcelChecker:
    """Excel文件检查器"""
    
    def __init__(self, config_file=CONFIG_FILE):
        """初始化检查器"""
        self.config = self._load_config(config_file)
        self.cache = self._load_cache()
        self.errors = []
        self.warnings = []
    
    def _load_config(self, config_file):
        """加载配置文件"""
        if os.path.exists(config_file):
            with open(config_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        return DEFAULT_CONFIG
    
    def _load_cache(self):
        """加载缓存"""
        if os.path.exists(CACHE_FILE):
            with open(CACHE_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        return {}
    
    def _save_cache(self):
        """保存缓存"""
        with open(CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump(self.cache, f, ensure_ascii=False, indent=2)
    
    def _calculate_file_hash(self, filepath):
        """计算文件哈希值"""
        hash_md5 = hashlib.md5()
        with open(filepath, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_md5.update(chunk)
        return hash_md5.hexdigest()
    
    def _get_remote_file_content(self, filepath, branch="main"):
        """获取远程仓库中的文件内容"""
        try:
            result = subprocess.run(
                ['git', 'show', f'{branch}:{filepath}'],
                capture_output=True,
                timeout=self.config['timeout']
            )
            
            if result.returncode == 0:
                return result.stdout, None
            else:
                return None, f"无法获取远程文件: {filepath}"
                
        except subprocess.TimeoutExpired:
            return None, f"获取远程文件超时: {filepath}"
        except Exception as e:
            return None, f"获取远程文件失败: {str(e)}"
    
    def _get_revision_records(self, filepath):
        """获取修订记录"""
        try:
            wb = load_workbook(filepath, read_only=True, data_only=True)
            
            # 检查是否存在修改记录sheet
            if RECORD_SHEET_NAME not in wb.sheetnames:
                wb.close()
                return None, "文件中不存在'修改记录'sheet页"
            
            ws = wb[RECORD_SHEET_NAME]
            records = []
            
            # 读取修订记录（从第2行开始，第1行是表头）
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and any(cell is not None for cell in row):
                    records.append({
                        "修订人": row[0] if len(row) > 0 else "",
                        "修订时间": row[1] if len(row) > 1 else "",
                        "修订内容": row[2] if len(row) > 2 else "",
                        "修订版本": row[3] if len(row) > 3 else ""
                    })
            
            wb.close()
            return records, None
            
        except Exception as e:
            return None, f"读取修订记录失败: {str(e)}"
    
    def _get_revision_records_from_bytes(self, content):
        """从字节内容获取修订记录"""
        try:
            # 创建临时文件
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
                tmp_file.write(content)
                tmp_file_path = tmp_file.name
            
            # 读取修订记录
            records, error = self._get_revision_records(tmp_file_path)
            
            # 删除临时文件
            os.unlink(tmp_file_path)
            
            return records, error
            
        except Exception as e:
            return None, f"从字节内容读取修订记录失败: {str(e)}"
    
    def _compare_revision_records(self, local_records, remote_records):
        """比较本地和远程的修订记录"""
        if not remote_records:
            return None, "远程文件没有修订记录，无法比较"
        
        if not local_records:
            return False, "本地文件没有修订记录"
        
        # 获取远程最新的修订记录
        remote_latest = remote_records[-1]
        
        # 检查本地是否包含远程最新的修订记录
        for record in local_records:
            if (record["修订人"] == remote_latest["修订人"] and
                record["修订时间"] == remote_latest["修订时间"] and
                record["修订内容"] == remote_latest["修订内容"]):
                return True, None
        
        # 本地不包含远程最新的修订记录
        return False, f"本地文件未包含远程最新的修订记录: {remote_latest['修订人']} - {remote_latest['修订时间']}"
    
    def _check_single_file(self, filepath, relative_path):
        """检查单个文件"""
        result = {
            "filepath": relative_path,
            "status": "pass",
            "errors": [],
            "warnings": []
        }
        
        # 计算当前文件哈希
        current_hash = self._calculate_file_hash(filepath)
        
        # 检查缓存中是否有记录
        if relative_path in self.cache:
            cached_hash = self.cache[relative_path].get("hash", "")
            
            # 如果哈希值相同，说明文件未修改，跳过检查
            if cached_hash == current_hash:
                result["status"] = "skipped"
                return result
        
        # 获取本地修订记录
        local_records, error = self._get_revision_records(filepath)
        
        if error:
            result["status"] = "error"
            result["errors"].append(error)
            return result
        
        # 检查修订记录是否为空
        if not local_records:
            result["status"] = "error"
            result["errors"].append("修改记录sheet页为空，请添加修订记录后再提交")
            return result
        
        # 获取远程文件内容
        remote_content, error = self._get_remote_file_content(relative_path)
        
        if error:
            # 无法获取远程文件，可能是新文件或网络问题，跳过版本检查
            result["warnings"].append(f"无法获取远程文件: {error}")
        else:
            # 获取远程修订记录
            remote_records, error = self._get_revision_records_from_bytes(remote_content)
            
            if error:
                result["warnings"].append(f"无法读取远程修订记录: {error}")
            else:
                # 比较本地和远程的修订记录
                is_up_to_date, error = self._compare_revision_records(local_records, remote_records)
                
                if error:
                    result["status"] = "error"
                    result["errors"].append(error)
                    return result
                
                if not is_up_to_date:
                    result["status"] = "error"
                    result["errors"].append(
                        "本地文件未基于远程最新版本，请先执行 'git pull' 获取最新版本，"
                        "在此基础上进行修改后再提交"
                    )
                    return result
        
        # 更新缓存
        self.cache[relative_path] = {
            "hash": current_hash,
            "last_check": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "record_count": len(local_records)
        }
        
        return result
    
    def check_files(self, file_list=None):
        """检查文件列表"""
        if file_list is None:
            # 检查所有Excel文件
            file_list = []
            if os.path.exists(EXCEL_DIR):
                for filename in os.listdir(EXCEL_DIR):
                    if filename.endswith('.xlsx'):
                        filepath = os.path.join(EXCEL_DIR, filename)
                        file_list.append((filepath, filename))
        
        if not file_list:
            print("没有找到需要检查的Excel文件")
            return True
        
        print(f"开始检查 {len(file_list)} 个Excel文件...")
        print(f"使用 {self.config['max_threads']} 个线程并行处理")
        print("-" * 60)
        
        # 使用线程池并行检查
        results = []
        with ThreadPoolExecutor(max_workers=self.config['max_threads']) as executor:
            future_to_file = {
                executor.submit(self._check_single_file, filepath, relative_path): (filepath, relative_path)
                for filepath, relative_path in file_list
            }
            
            for future in as_completed(future_to_file):
                filepath, relative_path = future_to_file[future]
                try:
                    result = future.result()
                    results.append(result)
                    
                    if result["status"] == "pass":
                        print(f"[OK] {relative_path} - 检查通过")
                    elif result["status"] == "skipped":
                        print(f"[SKIP] {relative_path} - 未修改，跳过检查")
                    elif result["status"] == "error":
                        print(f"[ERROR] {relative_path} - 检查失败")
                        for error in result["errors"]:
                            print(f"  错误: {error}")
                            self.errors.append(f"{relative_path}: {error}")
                    
                    # 显示警告
                    for warning in result.get("warnings", []):
                        print(f"  警告: {warning}")
                        self.warnings.append(f"{relative_path}: {warning}")
                        
                except Exception as e:
                    error_msg = f"{relative_path}: 检查时发生异常 - {str(e)}"
                    print(f"[ERROR] {error_msg}")
                    self.errors.append(error_msg)
        
        # 保存缓存
        self._save_cache()
        
        # 输出统计信息
        print("-" * 60)
        passed = sum(1 for r in results if r["status"] == "pass")
        skipped = sum(1 for r in results if r["status"] == "skipped")
        failed = sum(1 for r in results if r["status"] == "error")
        
        print(f"检查完成: 通过 {passed} 个, 跳过 {skipped} 个, 失败 {failed} 个")
        
        return len(self.errors) == 0


def main():
    """主函数"""
    import argparse
    
    parser = argparse.ArgumentParser(description='Excel文件检查器')
    parser.add_argument('--all', action='store_true', help='检查所有Excel文件')
    parser.add_argument('--files', nargs='+', help='指定要检查的文件列表')
    
    args = parser.parse_args()
    
    checker = ExcelChecker()
    
    if args.files:
        # 检查指定的文件
        file_list = []
        for filepath in args.files:
            if os.path.exists(filepath):
                relative_path = os.path.basename(filepath)
                file_list.append((filepath, relative_path))
        success = checker.check_files(file_list)
    elif args.all:
        # 检查所有文件
        success = checker.check_files()
    else:
        # 检查暂存区的文件
        try:
            # 获取暂存区的Excel文件
            result = subprocess.run(
                ['git', 'diff', '--cached', '--name-only', '--diff-filter=ACM'],
                capture_output=True,
                text=True,
                encoding='utf-8'
            )
            
            if result.returncode == 0:
                staged_files = result.stdout.strip().split('\n')
                file_list = []
                for filepath in staged_files:
                    if filepath.endswith('.xlsx') and os.path.exists(filepath):
                        file_list.append((filepath, filepath))
                
                if file_list:
                    success = checker.check_files(file_list)
                else:
                    print("暂存区中没有Excel文件需要检查")
                    success = True
            else:
                print("无法获取暂存区文件，检查所有Excel文件")
                success = checker.check_files()
        except Exception as e:
            print(f"Git命令执行失败: {str(e)}，检查所有Excel文件")
            success = checker.check_files()
    
    # 返回状态码
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
